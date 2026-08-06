import base64
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.dependencies import get_current_user
from app.crud.teaching_tasks import get_or_create_active_term
from app.db.session import get_db_session
from app.models.curriculum import ExperimentCourse, ExperimentProject
from app.models.identity import StudentBusyBitmap, Teacher
from app.models.resources import TeacherProjectQualification
from app.models.scheduling import (
    ExperimentSession,
    ProjectDemand,
    ScheduleVersion,
    TeacherTimetableEntry,
    TeachingTask,
)

router = APIRouter(prefix="/teachers", tags=["教师"])


def _compute_current_week(term) -> int:
    from datetime import date as dt_date
    today = dt_date.today()
    if today < term.start_date: return 0
    if today > term.end_date: return term.total_weeks
    return min((today - term.start_date).days // 7 + 1, term.total_weeks)


async def _get_teacher(session: AsyncSession, login_name: str) -> Teacher:
    teacher = (await session.execute(
        select(Teacher).where(Teacher.employee_no == login_name.upper())
    )).scalar_one_or_none()
    if teacher is None:
        raise HTTPException(status_code=404, detail="教师信息不存在")
    return teacher


@router.get("/me/profile")
async def get_my_profile(
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """教师个人信息 + 教学任务概览。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    # 有资格教授的项目
    quals = (await session.execute(
        select(TeacherProjectQualification)
        .where(
            TeacherProjectQualification.teacher_id == teacher.id,
            TeacherProjectQualification.status == "ACTIVE",
        )
    )).scalars().all()

    qualified_project_ids = [q.project_id for q in quals]

    # 查找对应的教学任务
    tasks = []
    if qualified_project_ids:
        demands = (await session.execute(
            select(ProjectDemand)
            .options(
                selectinload(ProjectDemand.project),
                selectinload(ProjectDemand.task).selectinload(TeachingTask.course),
            )
            .where(
                ProjectDemand.project_id.in_(qualified_project_ids),
                TeachingTask.term_id == term.id,
            )
            .join(TeachingTask, TeachingTask.id == ProjectDemand.task_id)
        )).scalars().all()

        seen = set()
        for d in demands:
            if d.task_id in seen:
                continue
            seen.add(d.task_id)
            tasks.append({
                "task_id": str(d.task_id),
                "project_id": str(d.project_id),
                "task_code": d.task.task_code,
                "course_name": d.task.course.course_name,
                "course_code": d.task.course.course_code,
                "planned_student_count": d.task.planned_student_count,
                "week_start": d.task.week_start,
                "week_end": d.task.week_end,
            })

    # 已排课场次数
    session_count = (await session.execute(
        select(TeacherTimetableEntry).where(
            TeacherTimetableEntry.teacher_id == teacher.id,
            TeacherTimetableEntry.term_id == term.id,
        )
    )).scalars().all()

    return {
        "name": teacher.name,
        "employee_no": teacher.employee_no,
        "department": teacher.department,
        "title": teacher.title,
        "term": {
            "academic_year": term.academic_year,
            "semester_no": term.semester_no,
            "start_date": str(term.start_date),
            "total_weeks": term.total_weeks,
            "current_week": _compute_current_week(term),
        },
        "qualified_projects_count": len(qualified_project_ids),
        "teaching_tasks": tasks,
        "scheduled_session_count": len(session_count),
    }


@router.get("/me/timetable")
async def get_my_timetable(
    week: int = Query(1, ge=0, le=20),
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """教师课表（week=0 返回全部周）。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    try:
        entries = (await session.execute(
            select(TeacherTimetableEntry)
            .where(
                TeacherTimetableEntry.teacher_id == teacher.id,
                TeacherTimetableEntry.term_id == term.id,
            )
        )).scalars().all()
    except Exception:
        entries = []

    if not entries:
        return {"week": week, "sessions": [], "total": 0}

    session_ids = [e.experiment_session_id for e in entries]
    where_clause = [ExperimentSession.id.in_(session_ids)]
    if week > 0:
        where_clause.append(ExperimentSession.week_no == week)

    es_list = (await session.execute(
        select(ExperimentSession)
        .options(
            selectinload(ExperimentSession.project),
            selectinload(ExperimentSession.laboratory),
        )
        .where(*where_clause)
        .order_by(ExperimentSession.week_no, ExperimentSession.day_of_week, ExperimentSession.start_slot)
    )).scalars().all()

    result = []
    for s in es_list:
        course = None
        if s.project:
            course = (await session.execute(
                select(ExperimentCourse).where(ExperimentCourse.id == s.project.course_id)
            )).scalar_one_or_none()

        result.append({
            "id": str(s.id),
            "day_of_week": s.day_of_week,
            "start_slot": s.start_slot,
            "end_slot": s.end_slot,
            "week_no": s.week_no,
            "capacity": s.capacity,
            "selected_count": s.selected_count,
            "project_name": s.project.project_name if s.project else "",
            "course_name": course.course_name if course else "",
            "course_code": course.course_code if course else "",
            "lab_name": s.laboratory.name if s.laboratory else "",
        })
    return {"week": week, "sessions": result, "total": len(result)}


@router.get("/me/upcoming")
async def get_upcoming_sessions(
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """获取教师最近的即将授课场次（按日期排序，取前5条）。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)
    cw = _compute_current_week(term)

    try:
        entries = (await session.execute(
            select(TeacherTimetableEntry).where(
                TeacherTimetableEntry.teacher_id == teacher.id,
                TeacherTimetableEntry.term_id == term.id,
            )
        )).scalars().all()
    except Exception:
        entries = []

    if not entries:
        return {"sessions": [], "current_week": cw}

    session_ids = [e.experiment_session_id for e in entries]
    # 取当前周及之后的场次
    es_list = (await session.execute(
        select(ExperimentSession)
        .options(
            selectinload(ExperimentSession.project),
            selectinload(ExperimentSession.laboratory),
        )
        .where(
            ExperimentSession.id.in_(session_ids),
            ExperimentSession.week_no >= max(1, cw),
        )
        .order_by(ExperimentSession.week_no, ExperimentSession.day_of_week, ExperimentSession.start_slot)
        .limit(5)
    )).scalars().all()

    result = []
    for s in es_list:
        course = None
        if s.project:
            course = (await session.execute(
                select(ExperimentCourse).where(ExperimentCourse.id == s.project.course_id)
            )).scalar_one_or_none()
        result.append({
            "id": str(s.id),
            "week_no": s.week_no,
            "day_of_week": s.day_of_week,
            "start_slot": s.start_slot,
            "end_slot": s.end_slot,
            "project_name": s.project.project_name if s.project else "",
            "course_name": course.course_name if course else "",
            "lab_name": s.laboratory.name if s.laboratory else "",
            "selected_count": s.selected_count,
            "capacity": s.capacity,
        })
    return {"sessions": result, "current_week": cw}


@router.get("/me/schedule/export")
async def export_schedule(
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """导出教师全部课表为 Excel 文件（每周一个 Sheet）。"""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    try:
        entries = (await session.execute(
            select(TeacherTimetableEntry).where(
                TeacherTimetableEntry.teacher_id == teacher.id,
                TeacherTimetableEntry.term_id == term.id,
            )
        )).scalars().all()
    except Exception:
        entries = []

    session_ids = [e.experiment_session_id for e in entries] if entries else []
    all_sessions = []
    if session_ids:
        all_sessions = (await session.execute(
            select(ExperimentSession)
            .options(selectinload(ExperimentSession.project), selectinload(ExperimentSession.laboratory))
            .where(ExperimentSession.id.in_(session_ids))
            .order_by(ExperimentSession.week_no, ExperimentSession.day_of_week, ExperimentSession.start_slot)
        )).scalars().all()

    # 按周分组
    weeks_map: dict[int, list] = {}
    for s in all_sessions:
        weeks_map.setdefault(s.week_no, []).append(s)

    wb = Workbook()
    wb.remove(wb.active)

    days = ["周日", "周一", "周二", "周三", "周四", "周五", "周六"]
    header_font = Font(bold=True)

    total_weeks = max(weeks_map.keys()) if weeks_map else 18
    for wk in range(1, total_weeks + 1):
        ws = wb.create_sheet(title=f"第{wk}周")
        for col, d in enumerate(days, 2):
            ws.cell(row=1, column=col, value=d).font = header_font
        for row in range(2, 14):
            ws.cell(row=row, column=1, value=f"第{row-1}节")

        for s in weeks_map.get(wk, []):
            course = None
            if s.project:
                course = (await session.execute(
                    select(ExperimentCourse).where(ExperimentCourse.id == s.project.course_id)
                )).scalar_one_or_none()
            text = f"{s.project.project_name if s.project else ''}\n{getattr(course, 'course_name', '')}\n{s.laboratory.name if s.laboratory else ''}"
            # DB day_of_week: 1=Mon..7=Sun → Excel col: 2=Sun(B)..8=Sat(H)
            col = 2 if s.day_of_week == 7 else s.day_of_week + 2
            start_row = s.start_slot + 1
            end_row = s.end_slot + 1
            cell = ws.cell(row=start_row, column=col, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        ws.column_dimensions['A'].width = 8
        for c in range(2, 9):
            ws.column_dimensions[chr(64 + c)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = "jskb.xlsx"  # teacher-external name set by frontend
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@router.get("/me/busy-bitmap")
async def get_my_bitmap(
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """教师忙闲位图（从排课结果生成）。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    try:
        from app.models.identity import TeacherBusyBitmap
        bm = (await session.execute(
            select(TeacherBusyBitmap).where(
                TeacherBusyBitmap.teacher_id == teacher.id,
                TeacherBusyBitmap.term_id == term.id,
            )
        )).scalar_one_or_none()
    except Exception:
        return {"weeks": 18, "days": 7, "slots": 12, "data": None}

    if bm is None:
        return {"weeks": 18, "days": 7, "slots": 12, "data": None}

    return {
        "weeks": bm.end_week - bm.start_week + 1,
        "days": bm.days_per_week,
        "slots": bm.slots_per_day,
        "data": base64.b64encode(bm.bitmap).decode(),
        "start_week": bm.start_week,
        "end_week": bm.end_week,
    }


@router.get("/me/projects")
async def get_my_projects(
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """教师授课项目列表（每个项目含所有场次）。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    entries = (await session.execute(
        select(TeacherTimetableEntry).where(
            TeacherTimetableEntry.teacher_id == teacher.id,
            TeacherTimetableEntry.term_id == term.id,
        )
    )).scalars().all()

    if not entries:
        return {"projects": []}

    session_ids = [e.experiment_session_id for e in entries]
    es_list = (await session.execute(
        select(ExperimentSession)
        .options(
            selectinload(ExperimentSession.project),
            selectinload(ExperimentSession.laboratory),
        )
        .where(ExperimentSession.id.in_(session_ids))
        .order_by(ExperimentSession.project_id, ExperimentSession.week_no, ExperimentSession.day_of_week)
    )).scalars().all()

    # 按 project_id 分组
    project_map: dict = {}
    for s in es_list:
        pid = str(s.project_id)
        if pid not in project_map:
            project_map[pid] = {
                "project_id": pid,
                "project_name": s.project.project_name if s.project else "",
                "sessions": [],
            }
        project_map[pid]["sessions"].append({
            "session_id": str(s.id),
            "week_no": s.week_no,
            "day_of_week": s.day_of_week,
            "start_slot": s.start_slot,
            "end_slot": s.end_slot,
            "lab_name": s.laboratory.name if s.laboratory else "",
            "capacity": s.capacity,
            "selected_count": s.selected_count,
        })

    return {"projects": list(project_map.values())}


@router.get("/me/session-students")
async def get_session_students(
    session_id: UUID = Query(...),
    db: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """查某个场次已选课的学生列表。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    from app.models.enrollment import StudentProjectRecord
    from app.models.identity import Student as StudentModel, Major as MajorModel, UserAccount

    records = (await db.execute(
        select(StudentProjectRecord, StudentModel)
        .join(StudentModel, StudentModel.id == StudentProjectRecord.student_id)
        .where(
            StudentProjectRecord.session_id == session_id,
            StudentProjectRecord.status == "SELECTED",
        )
    )).all()

    student_ids = [r[1].id for r in records]
    user_ids = [r[1].user_id for r in records]
    major_map = {}
    phone_map = {}
    if student_ids:
        rows = (await db.execute(
            select(StudentModel.id, MajorModel.name)
            .join(MajorModel, MajorModel.id == StudentModel.major_id)
            .where(StudentModel.id.in_(student_ids))
        )).all()
        major_map = {r[0]: r[1] for r in rows}
        phones = (await db.execute(
            select(UserAccount.id, UserAccount.phone)
            .where(UserAccount.id.in_(user_ids))
        )).all()
        phone_map = {r[0]: r[1] for r in phones}

    result = []
    seen = set()
    for record, student in records:
        if student.id in seen:
            continue
        seen.add(student.id)
        phone = phone_map.get(student.user_id) or ''
        result.append({
            "student_id": str(student.id),
            "name": student.name,
            "student_no": student.student_no,
            "major_name": major_map.get(student.id, ""),
            "enrollment_year": student.enrollment_year,
            "phone": phone or '',
        })

    return {"session_id": str(session_id), "students": result}


@router.get("/me/project-students")
async def get_project_students(
    project_id: UUID = Query(...),
    session: AsyncSession = Depends(get_db_session),
    current_user=Depends(get_current_user),
):
    """查某项目下已选课的学生列表。"""
    if current_user.user_type != "TEACHER":
        raise HTTPException(status_code=403, detail="仅教师可查询")

    teacher = await _get_teacher(session, current_user.login_name)
    term = await get_or_create_active_term(session)

    # 找到该教师在该项目下的所有场次
    sessions = (await session.execute(
        select(ExperimentSession.id).where(
            ExperimentSession.teacher_id == teacher.id,
            ExperimentSession.project_id == project_id,
        )
    )).scalars().all()

    if not sessions:
        return {"project_id": str(project_id), "students": []}

    # 查选了这些场次的学生记录
    from app.models.enrollment import StudentProjectRecord
    from app.models.identity import Student as StudentModel, Major as MajorModel

    records = (await session.execute(
        select(StudentProjectRecord, StudentModel)
        .join(StudentModel, StudentModel.id == StudentProjectRecord.student_id)
        .where(
            StudentProjectRecord.session_id.in_(sessions),
            StudentProjectRecord.status == "SELECTED",
        )
        .order_by(StudentProjectRecord.selected_at.desc())
    )).all()

    # 批量查 major
    student_ids = [r[1].id for r in records]
    major_map = {}
    if student_ids:
        majors = (await session.execute(
            select(StudentModel.id, MajorModel.name)
            .join(MajorModel, MajorModel.id == StudentModel.major_id)
            .where(StudentModel.id.in_(student_ids))
        )).all()
        major_map = {m[0]: m[1] for m in majors}

    result = []
    seen = set()
    for r in records:
        stu = r[1]
        if stu.id in seen:
            continue
        seen.add(stu.id)
        result.append({
            "student_id": str(stu.id),
            "name": stu.name,
            "student_no": stu.student_no,
            "major_name": major_map.get(stu.id, ""),
            "enrollment_year": stu.enrollment_year,
        })

    return {"project_id": str(project_id), "students": result}
